import openpyxl
import json

# Path to Excel file
excel_path = r'C:\Users\GiorgosKorifidis\Downloads\bonus_template_Black+Friday_+Casino+Reload+200%+up+to+€300+21.11.25\Book1.xlsx'

try:
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Get headers
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)

    print("=== EXCEL HEADERS ===")
    for i, header in enumerate(headers, 1):
        print(f"{i}. {header}")

    print("\n=== FIRST 5 ROWS OF DATA ===")
    for row_idx in range(2, min(7, ws.max_row + 1)):
        row_data = {}
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            row_data[header] = cell.value
        print(json.dumps(row_data, indent=2, default=str))
        print("---")

except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
